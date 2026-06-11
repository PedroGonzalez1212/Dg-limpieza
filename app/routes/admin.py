from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify
from flask_login import login_required, current_user
from app import db
from app.models import Product, Category, ProductVariant, StockMovement, Sale, SaleItem, User, Combo, ComboItem
from app.utils.decorators import admin_required
from datetime import datetime, date, timedelta

import cloudinary
import cloudinary.uploader
import os

cloudinary.config(
    cloud_name = os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key    = os.environ.get('CLOUDINARY_API_KEY'),
    api_secret = os.environ.get('CLOUDINARY_API_SECRET'),
    secure     = True
)

admin_bp = Blueprint('admin', __name__)


# ── Helpers ──────────────────────────────────────────────────────────────────
def _precio_float(valor):
    """Convierte un string de formulario a float, o None si está vacío."""
    try:
        return float(valor) if valor and valor.strip() else None
    except (ValueError, TypeError):
        return None

def _int_or_none(valor):
    """Convierte un string de formulario a int, o None si está vacío."""
    try:
        return int(valor) if valor and valor.strip() else None
    except (ValueError, TypeError):
        return None


# ════════════════════════════════════════════════════════════════════════════
# LISTADO DE PRODUCTOS
# ════════════════════════════════════════════════════════════════════════════
@admin_bp.route('/productos')
@login_required
@admin_required
def productos():
    """
    Lista todos los productos (activos e inactivos) con sus datos clave.
    Permite buscar por nombre y filtrar por categoría.
    """
    busqueda   = request.args.get('q', '').strip()
    categoria  = request.args.get('categoria', '', type=str)
    mostrar    = request.args.get('mostrar', 'activos')  # 'activos', 'inactivos', 'todos'

    query = Product.query

    # Filtro de estado
    if mostrar == 'activos':
        query = query.filter_by(activo=True)
    elif mostrar == 'inactivos':
        query = query.filter_by(activo=False)
    elif mostrar == 'sin_stock':
        query = query.filter(Product.activo == True, Product.stock == 0)
    elif mostrar == 'stock_bajo':
        query = query.filter(
            Product.activo == True,
            Product.stock > 0,
            Product.stock <= Product.stock_minimo
        )
    # 'todos' → sin filtro

    # Búsqueda por nombre (ilike = case-insensitive)
    if busqueda:
        query = query.filter(Product.nombre.ilike(f'%{busqueda}%'))

    # Filtro por categoría
    if categoria:
        query = query.join(Product.categorias).filter(Category.slug == categoria)

    productos_list = query.order_by(Product.nombre).all()
    categorias     = Category.query.filter_by(activa=True).all()

    # Contadores para los badges del header
    total_activos   = Product.query.filter_by(activo=True).count()
    total_inactivos = Product.query.filter_by(activo=False).count()
    stock_bajo      = Product.query.filter(
        Product.activo == True,
        Product.stock <= Product.stock_minimo,
        Product.stock > 0
    ).count()
    sin_stock = Product.query.filter(
        Product.activo == True,
        Product.stock == 0
    ).count()

    return render_template('admin/productos.html',
        productos=productos_list,
        categorias=categorias,
        busqueda=busqueda,
        categoria_sel=categoria,
        mostrar=mostrar,
        total_activos=total_activos,
        total_inactivos=total_inactivos,
        stock_bajo=stock_bajo,
        sin_stock=sin_stock
    )


# ════════════════════════════════════════════════════════════════════════════
# NUEVO PRODUCTO
# ════════════════════════════════════════════════════════════════════════════
@admin_bp.route('/productos/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def producto_nuevo():
    categorias = Category.query.filter_by(activa=True).all()

    if request.method == 'GET':
        return render_template('admin/producto_form.html',
            producto=None,
            categorias=categorias,
            accion='nuevo'
        )

    # ── POST: procesar formulario ─────────────────────────────────────────
    nombre = request.form.get('nombre', '').strip()
    if not nombre:
        flash('El nombre del producto es obligatorio.', 'error')
        return render_template('admin/producto_form.html',
            producto=None, categorias=categorias, accion='nuevo')

    producto = Product(
        nombre            = nombre,
        descripcion       = request.form.get('descripcion', '').strip() or None,
        imagen_url        = request.form.get('imagen_url', '').strip() or None,
        precio_unidad     = _precio_float(request.form.get('precio_unidad')),
        precio_mayor      = _precio_float(request.form.get('precio_mayor')),
        precio_caja       = _precio_float(request.form.get('precio_caja')),
        cantidad_mayor    = _int_or_none(request.form.get('cantidad_mayor')),
        unidades_por_caja = _int_or_none(request.form.get('unidades_por_caja')),
        stock             = int(request.form.get('stock', 0) or 0),
        stock_minimo      = int(request.form.get('stock_minimo', 5) or 5),
        tiene_variantes   = 'tiene_variantes' in request.form,
        activo            = True
    )

    # Validación mínima de precio
    if not producto.precio_unidad:
        flash('El precio por unidad es obligatorio.', 'error')
        return render_template('admin/producto_form.html',
            producto=None, categorias=categorias, accion='nuevo')

    # Asociar categorías (checkbox múltiple)
    cats_ids = request.form.getlist('categorias')  # lista de IDs como strings
    for cat_id in cats_ids:
        cat = Category.query.get(int(cat_id))
        if cat:
            producto.categorias.append(cat)

    db.session.add(producto)
    db.session.flush()  # flush para obtener producto.id antes del commit

    # ── Variantes (si tiene) ──────────────────────────────────────────────
    # El formulario envía listas paralelas: variante_nombre[], variante_valor[], etc.
    if producto.tiene_variantes:
        nombres  = request.form.getlist('variante_nombre')
        valores  = request.form.getlist('variante_valor')
        stocks   = request.form.getlist('variante_stock')
        extras   = request.form.getlist('variante_precio_extra')

        for i, (n, v) in enumerate(zip(nombres, valores)):
            if n.strip() and v.strip():
                variante = ProductVariant(
                    producto_id  = producto.id,
                    nombre       = n.strip(),
                    valor        = v.strip(),
                    stock        = int(stocks[i]) if i < len(stocks) and stocks[i] else 0,
                    precio_extra = _precio_float(extras[i]) if i < len(extras) else 0
                )
                db.session.add(variante)

    # Registrar movimiento de stock inicial
    if producto.stock > 0:
        mov = StockMovement(
            producto_id = producto.id,
            usuario_id  = current_user.id,
            tipo        = 'entrada',
            cantidad    = producto.stock,
            motivo      = 'Stock inicial al crear producto'
        )
        db.session.add(mov)

    db.session.commit()
    flash(f'Producto "{producto.nombre}" creado correctamente.', 'success')
    return redirect(url_for('admin.productos'))


# ════════════════════════════════════════════════════════════════════════════
# EDITAR PRODUCTO
# ════════════════════════════════════════════════════════════════════════════
@admin_bp.route('/productos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def producto_editar(id):
    producto   = Product.query.get_or_404(id)
    categorias = Category.query.filter_by(activa=True).all()

    if request.method == 'GET':
        return render_template('admin/producto_form.html',
            producto=producto,
            categorias=categorias,
            accion='editar'
        )

    # ── POST: guardar cambios ─────────────────────────────────────────────
    nombre = request.form.get('nombre', '').strip()
    if not nombre:
        flash('El nombre del producto es obligatorio.', 'error')
        return render_template('admin/producto_form.html',
            producto=producto, categorias=categorias, accion='editar')

    stock_anterior = producto.stock

    producto.nombre            = nombre
    producto.descripcion       = request.form.get('descripcion', '').strip() or None
    producto.imagen_url        = request.form.get('imagen_url', '').strip() or None
    producto.precio_unidad     = _precio_float(request.form.get('precio_unidad'))
    producto.precio_mayor      = _precio_float(request.form.get('precio_mayor'))
    producto.precio_caja       = _precio_float(request.form.get('precio_caja'))
    producto.cantidad_mayor    = _int_or_none(request.form.get('cantidad_mayor'))
    producto.unidades_por_caja = _int_or_none(request.form.get('unidades_por_caja'))
    producto.stock             = int(request.form.get('stock', 0) or 0)
    producto.stock_minimo      = int(request.form.get('stock_minimo', 5) or 5)
    producto.tiene_variantes   = 'tiene_variantes' in request.form

    # Recalcular categorías: limpiar y reasignar
    producto.categorias.clear()
    for cat_id in request.form.getlist('categorias'):
        cat = Category.query.get(int(cat_id))
        if cat:
            producto.categorias.append(cat)

    # Registrar movimiento si cambió el stock
    diferencia = producto.stock - stock_anterior
    if diferencia != 0:
        mov = StockMovement(
            producto_id = producto.id,
            usuario_id  = current_user.id,
            tipo        = 'entrada' if diferencia > 0 else 'ajuste',
            cantidad    = abs(diferencia),
            motivo      = f'Ajuste manual desde edición de producto ({"+" if diferencia > 0 else ""}{diferencia})'
        )
        db.session.add(mov)

    db.session.commit()
    flash(f'Producto "{producto.nombre}" actualizado.', 'success')
    return redirect(url_for('admin.productos'))


# ════════════════════════════════════════════════════════════════════════════
# DESACTIVAR / REACTIVAR PRODUCTO
# ════════════════════════════════════════════════════════════════════════════
@admin_bp.route('/productos/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def producto_toggle(id):
    """
    Activa o desactiva un producto (no lo borra).
    Recibe POST desde un formulario con botón.
    """
    producto = Product.query.get_or_404(id)
    producto.activo = not producto.activo
    db.session.commit()

    estado = 'activado' if producto.activo else 'desactivado'
    flash(f'Producto "{producto.nombre}" {estado}.', 'success')
    return redirect(url_for('admin.productos'))


# ════════════════════════════════════════════════════════════════════════════
# DETALLE / VISTA RÁPIDA (JSON — para modal opcional)
# ════════════════════════════════════════════════════════════════════════════
@admin_bp.route('/productos/<int:id>/json')
@login_required
@admin_required
def producto_json(id):
    """Endpoint auxiliar que devuelve los datos del producto en JSON."""
    p = Product.query.get_or_404(id)
    return jsonify({
        'id':             p.id,
        'nombre':         p.nombre,
        'stock':          p.stock,
        'stock_minimo':   p.stock_minimo,
        'precio_unidad':  float(p.precio_unidad),
        'activo':         p.activo,
    })
    
# ════════════════════════════════════════════════════════════════════════════
# GESTIÓN DE STOCK
# Agregá estas rutas al final de tu app/routes/admin.py
# ════════════════════════════════════════════════════════════════════════════

@admin_bp.route('/stock')
@login_required
@admin_required
def stock():
    """
    Pantalla principal de gestión de stock.
    Muestra todos los productos activos con su nivel de stock actual.
    Permite filtrar por estado de stock para trabajar por prioridad.
    """
    filtro = request.args.get('filtro', 'todos')  # 'todos', 'criticos', 'sin_stock'

    query = Product.query.filter_by(activo=True)

    if filtro == 'sin_stock':
        query = query.filter(Product.stock == 0)
    elif filtro == 'criticos':
        # Stock bajo: mayor que 0 pero igual o menor al mínimo
        query = query.filter(
            Product.stock > 0,
            Product.stock <= Product.stock_minimo
        )

    productos_list = query.order_by(Product.stock.asc(), Product.nombre).all()

    # Contadores para los tabs del filtro
    total      = Product.query.filter_by(activo=True).count()
    sin_stock  = Product.query.filter(Product.activo == True, Product.stock == 0).count()
    criticos   = Product.query.filter(
        Product.activo == True,
        Product.stock > 0,
        Product.stock <= Product.stock_minimo
    ).count()

    return render_template('admin/stock.html',
        productos=productos_list,
        filtro=filtro,
        total=total,
        sin_stock=sin_stock,
        criticos=criticos
    )


@admin_bp.route('/stock/<int:id>/ajustar', methods=['POST'])
@login_required
@admin_required
def stock_ajustar(id):
    """
    Procesa un ajuste de stock para un producto.
    
    Recibe por POST:
    - tipo: 'entrada' (mercadería nueva) o 'ajuste' (corrección manual)
    - cantidad: número entero positivo
    - motivo: texto libre opcional
    
    Para 'entrada' → suma al stock actual.
    Para 'ajuste'  → reemplaza el stock actual por la cantidad indicada.
    
    ¿Por qué esta distinción? Porque "entró mercadería" y "corrijo el stock"
    son dos intenciones distintas con diferentes registros contables.
    Una entrada suma; un ajuste establece el valor correcto.
    """
    producto = Product.query.get_or_404(id)

    tipo     = request.form.get('tipo', 'entrada')
    cantidad = request.form.get('cantidad', '0').strip()
    motivo   = request.form.get('motivo', '').strip() or None

    # Validar que la cantidad sea un número entero positivo
    try:
        cantidad = int(cantidad)
        if cantidad < 0:
            raise ValueError
    except (ValueError, TypeError):
        flash('La cantidad debe ser un número entero positivo.', 'error')
        return redirect(url_for('admin.stock'))

    stock_anterior = producto.stock

    if tipo == 'entrada':
        # Suma al stock existente
        producto.stock += cantidad
        cantidad_mov = cantidad
    elif tipo == 'ajuste':
        # Reemplaza el stock. Guardamos la diferencia en el movimiento
        # para que el historial muestre cuánto cambió realmente.
        diferencia = cantidad - stock_anterior
        producto.stock = cantidad
        cantidad_mov = abs(diferencia)
        # Si no cambió nada, no tiene sentido registrar el movimiento
        if diferencia == 0:
            flash(f'El stock de "{producto.nombre}" ya era {cantidad}. Sin cambios.', 'info')
            return redirect(url_for('admin.stock'))
        # Si el ajuste fue a la baja, lo marcamos como ajuste de salida
        if diferencia < 0:
            tipo = 'ajuste'  # tipo ya es 'ajuste', pero dejamos explícito
    else:
        flash('Tipo de movimiento inválido.', 'error')
        return redirect(url_for('admin.stock'))

    # Registrar el movimiento en el historial
    mov = StockMovement(
        producto_id = producto.id,
        usuario_id  = current_user.id,
        tipo        = tipo,
        cantidad    = cantidad_mov,
        motivo      = motivo or f'Ajuste manual por {current_user.nombre}'
    )
    db.session.add(mov)
    db.session.commit()

    flash(
        f'Stock de "{producto.nombre}" actualizado: {stock_anterior} → {producto.stock}.',
        'success'
    )
    return redirect(url_for('admin.stock'))


@admin_bp.route('/stock/<int:id>/historial')
@login_required
@admin_required
def stock_historial(id):
    """
    Devuelve el historial de movimientos de un producto en JSON.
    Se consume desde el modal de historial via fetch().
    
    Limitamos a los últimos 30 movimientos para no sobrecargar el modal.
    Si el admin necesita más, en el Sprint 5 lo mostramos en los reportes.
    """
    producto = Product.query.get_or_404(id)

    movimientos = (
        StockMovement.query
        .filter_by(producto_id=id)
        .order_by(StockMovement.creado_en.desc())
        .limit(30)
        .all()
    )

    data = {
        'producto': producto.nombre,
        'stock_actual': producto.stock,
        'movimientos': [
            {
                'id':        m.id,
                'tipo':      m.tipo,
                'cantidad':  m.cantidad,
                'motivo':    m.motivo or '—',
                'usuario':   m.usuario_id,  # lo resolvemos abajo
                'fecha':     m.creado_en.strftime('%d/%m/%Y %H:%M')
            }
            for m in movimientos
        ]
    }

    # Resolver nombres de usuarios en un solo query (evita N+1 queries)
    from app.models import User
    user_ids  = {m.usuario_id for m in movimientos}
    usuarios  = {u.id: u.nombre for u in User.query.filter(User.id.in_(user_ids)).all()}
    for item in data['movimientos']:
        item['usuario'] = usuarios.get(item['usuario'], 'Sistema')

    return jsonify(data)

# ════════════════════════════════════════════════════════════════════════════
# HISTORIAL DE VENTAS — agregá estas rutas al final de app/routes/admin.py
# ════════════════════════════════════════════════════════════════════════════
# También agregá Sale y SaleItem al import que ya tenés arriba:
#
#   from app.models import Product, Category, ProductVariant, StockMovement, Sale, SaleItem
#
# ════════════════════════════════════════════════════════════════════════════

from datetime import date, timedelta  # solo si no lo tenés importado ya


@admin_bp.route('/ventas')
@login_required
@admin_required
def ventas():
    """
    Listado de todas las ventas con filtros.
    
    Filtros disponibles:
    - desde / hasta : rango de fechas (string YYYY-MM-DD)
    - tipo          : 'pos', 'whatsapp' o '' (todos)
    - metodo_pago   : cualquier string del campo, o '' (todos)
    - q             : búsqueda por nombre o teléfono del cliente

    ¿Por qué filtramos por fecha en el servidor y no con JS?
    Porque la tabla puede tener miles de registros. Filtrar en DB es O(log n)
    con un índice; filtrar en el navegador requiere bajar todos los datos primero.
    """
    # ── Parámetros de filtro ─────────────────────────────────────────────
    desde_str    = request.args.get('desde', '')
    hasta_str    = request.args.get('hasta', '')
    tipo         = request.args.get('tipo', '')
    metodo_pago  = request.args.get('metodo_pago', '')
    busqueda     = request.args.get('q', '').strip()

    # Si no hay fechas, mostrar los últimos 30 días por defecto
    hoy   = date.today()
    desde = date.fromisoformat(desde_str) if desde_str else hoy - timedelta(days=30)
    hasta = date.fromisoformat(hasta_str) if hasta_str else hoy

    # ── Query base ───────────────────────────────────────────────────────
    query = Sale.query.filter(
        Sale.creado_en >= datetime.combine(desde, datetime.min.time()),
        Sale.creado_en <= datetime.combine(hasta, datetime.max.time())
    )

    if tipo:
        query = query.filter_by(tipo=tipo)

    if metodo_pago:
        query = query.filter_by(metodo_pago=metodo_pago)

    if busqueda:
        like = f'%{busqueda}%'
        query = query.filter(
            db.or_(
                Sale.nombre_cliente.ilike(like),
                Sale.telefono_cliente.ilike(like)
            )
        )

    ventas_list = query.order_by(Sale.creado_en.desc()).all()

    # ── Totales para el resumen del encabezado ───────────────────────────
    # Calculamos sobre los resultados filtrados (no toda la tabla)
    total_ventas   = len(ventas_list)
    monto_total    = sum(float(v.total) for v in ventas_list)

    # Métodos de pago únicos que existen en la DB (para el select del filtro)
    metodos_disponibles = (
        db.session.query(Sale.metodo_pago)
        .filter(Sale.metodo_pago.isnot(None))
        .distinct()
        .all()
    )
    metodos_disponibles = [m[0] for m in metodos_disponibles if m[0]]

    return render_template('admin/ventas.html',
        ventas=ventas_list,
        desde=desde.isoformat(),
        hasta=hasta.isoformat(),
        tipo=tipo,
        metodo_pago=metodo_pago,
        busqueda=busqueda,
        total_ventas=total_ventas,
        monto_total=monto_total,
        metodos_disponibles=metodos_disponibles,
    )


@admin_bp.route('/ventas/<int:id>')
@login_required
@admin_required
def venta_detalle(id):
    """
    Detalle completo de una venta: datos del cliente, items y totales.
    
    Usamos get_or_404 para que Flask devuelva un 404 automáticamente
    si el ID no existe, en vez de explotar con un error 500.
    """
    venta = Sale.query.get_or_404(id)

    # Los items ya vienen cargados por la relación `items` definida en el modelo.
    # SQLAlchemy hace el JOIN automáticamente cuando accedés a venta.items.

    return render_template('admin/venta_detalle.html', venta=venta)

# ════════════════════════════════════════════════════════════════════════════
# GESTIÓN DE USUARIOS
# ════════════════════════════════════════════════════════════════════════════

@admin_bp.route('/usuarios')
@login_required
@admin_required
def usuarios():
    """
    Lista todos los usuarios del sistema.
    ¿Por qué mostramos también los inactivos?
    Porque el admin necesita poder reactivarlos — si no los ve, no puede hacerlo.
    """
    todos = User.query.order_by(User.nombre).all()
    return render_template('admin/usuarios.html', usuarios=todos)


@admin_bp.route('/usuarios/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def usuario_nuevo():
    if request.method == 'POST':
        nombre    = request.form.get('nombre', '').strip()
        email     = request.form.get('email', '').strip().lower()
        password  = request.form.get('password', '')
        rol       = request.form.get('rol', 'vendedor')

        # Validaciones básicas
        if not nombre or not email or not password:
            flash('Completá todos los campos.', 'error')
            return redirect(url_for('admin.usuario_nuevo'))

        if User.query.filter_by(email=email).first():
            flash('Ya existe un usuario con ese email.', 'error')
            return redirect(url_for('admin.usuario_nuevo'))

        usuario = User(nombre=nombre, email=email, rol=rol, activo=True)
        usuario.set_password(password)
        db.session.add(usuario)
        db.session.commit()

        flash(f'Usuario {nombre} creado correctamente.', 'ok')
        return redirect(url_for('admin.usuarios'))

    return render_template('admin/usuario_form.html', usuario=None)


@admin_bp.route('/usuarios/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def usuario_editar(id):
    usuario = User.query.get_or_404(id)

    if request.method == 'POST':
        usuario.nombre = request.form.get('nombre', '').strip()
        usuario.email  = request.form.get('email', '').strip().lower()
        usuario.rol    = request.form.get('rol', 'vendedor')

        # La contraseña es opcional al editar — si viene vacía, no la tocamos
        nueva_password = request.form.get('password', '').strip()
        if nueva_password:
            usuario.set_password(nueva_password)

        # Evitar que el admin se cambie el email a uno ya existente
        duplicado = User.query.filter(
            User.email == usuario.email,
            User.id != id
        ).first()
        if duplicado:
            flash('Ese email ya está en uso por otro usuario.', 'error')
            return redirect(url_for('admin.usuario_editar', id=id))

        db.session.commit()
        flash('Usuario actualizado.', 'ok')
        return redirect(url_for('admin.usuarios'))

    return render_template('admin/usuario_form.html', usuario=usuario)


@admin_bp.route('/usuarios/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def usuario_toggle(id):
    """
    Activa o desactiva un usuario.
    Usamos POST y no GET porque esta acción modifica datos.
    Un GET podría ejecutarse sin querer (prefetch del navegador, bots, etc).
    Además protegemos que el admin no se desactive a sí mismo.
    """
    usuario = User.query.get_or_404(id)

    if usuario.id == current_user.id:
        flash('No podés desactivarte a vos mismo.', 'error')
        return redirect(url_for('admin.usuarios'))

    usuario.activo = not usuario.activo
    db.session.commit()

    estado = 'activado' if usuario.activo else 'desactivado'
    flash(f'Usuario {usuario.nombre} {estado}.', 'ok')
    return redirect(url_for('admin.usuarios'))

# ════════════════════════════════════════════════════════════
# SPRINT 5 — REPORTES Y EXPORTACIÓN
# ════════════════════════════════════════════════════════════
from io import BytesIO
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@admin_bp.route('/reportes')
@login_required
@admin_required
def reportes():
    """
    Pantalla de reportes con filtros y tabla resumen.
    Reutilizamos la misma lógica de filtros que en /ventas,
    pero acá el foco es el análisis agregado + botones de exportar.
    """
    desde_str = request.args.get('desde', '')
    hasta_str = request.args.get('hasta', '')
    metodo_pago = request.args.get('metodo_pago', '')

    # Fechas por defecto: mes actual
    hoy = date.today()
    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else hoy.replace(day=1)
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy
    except ValueError:
        desde = hoy.replace(day=1)
        hasta = hoy

    # Query base
    query = Sale.query.filter(
        Sale.creado_en >= datetime.combine(desde, datetime.min.time()),
        Sale.creado_en <= datetime.combine(hasta, datetime.max.time()),
    )
    if metodo_pago:
        query = query.filter(Sale.metodo_pago == metodo_pago)

    ventas = query.order_by(Sale.creado_en.desc()).all()

    # Métricas para el resumen
    total_ventas   = len(ventas)
    monto_total    = sum(float(v.total) for v in ventas)
    ticket_promedio = monto_total / total_ventas if total_ventas > 0 else 0

    # Agrupación por método de pago (para el mini-gráfico)
    from collections import defaultdict
    por_metodo = defaultdict(float)
    for v in ventas:
        clave = v.metodo_pago or 'Sin especificar'
        por_metodo[clave] += float(v.total)

    metodos_disponibles = db.session.query(Sale.metodo_pago)\
        .filter(Sale.metodo_pago != None)\
        .distinct().all()
    metodos_disponibles = [m[0] for m in metodos_disponibles]

    return render_template('admin/reportes.html',
        ventas=ventas,
        desde=desde.isoformat(),
        hasta=hasta.isoformat(),
        metodo_pago=metodo_pago,
        total_ventas=total_ventas,
        monto_total=monto_total,
        ticket_promedio=ticket_promedio,
        por_metodo=dict(por_metodo),
        metodos_disponibles=metodos_disponibles,
    )


@admin_bp.route('/reportes/exportar/excel')
@login_required
@admin_required
def exportar_excel():
    """
    Genera y devuelve un archivo .xlsx con las ventas filtradas.
    
    Usamos BytesIO (un buffer en memoria) en vez de guardar el archivo
    en disco. Así no acumulamos archivos temporales en el servidor.
    Flask lo sirve directamente como descarga con send_file().
    """
    from flask import send_file

    desde_str   = request.args.get('desde', '')
    hasta_str   = request.args.get('hasta', '')
    metodo_pago = request.args.get('metodo_pago', '')

    hoy = date.today()
    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else hoy.replace(day=1)
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy
    except ValueError:
        desde = hoy.replace(day=1)
        hasta = hoy

    query = Sale.query.filter(
        Sale.creado_en >= datetime.combine(desde, datetime.min.time()),
        Sale.creado_en <= datetime.combine(hasta, datetime.max.time()),
    )
    if metodo_pago:
        query = query.filter(Sale.metodo_pago == metodo_pago)
    ventas = query.order_by(Sale.creado_en.asc()).all()

    # ── Construir el Excel ───────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    # Colores de la paleta DG Limpieza
    COLOR_HEADER = '0C0C62'   # navy
    COLOR_FILA_PAR = 'EAF6FF'  # celeste muy claro

    # Encabezados
    encabezados = ['#', 'Fecha', 'Cliente', 'Teléfono', 'Canal', 'Método de pago', 'Total']
    for col, titulo in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=titulo)
        celda.font      = Font(bold=True, color='FFFFFF', size=11)
        celda.fill      = PatternFill('solid', fgColor=COLOR_HEADER)
        celda.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 22

    # Filas de datos
    for fila, venta in enumerate(ventas, 2):
        datos = [
            venta.id,
            venta.creado_en.strftime('%d/%m/%Y %H:%M'),
            venta.nombre_cliente or '—',
            venta.telefono_cliente or '—',
            venta.tipo.upper(),
            venta.metodo_pago or '—',
            float(venta.total),
        ]
        for col, valor in enumerate(datos, 1):
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.alignment = Alignment(horizontal='center')
            # Filas alternadas para legibilidad
            if fila % 2 == 0:
                celda.fill = PatternFill('solid', fgColor=COLOR_FILA_PAR)

        # Columna Total: formato de moneda
        ws.cell(row=fila, column=7).number_format = '"$"#,##0.00'

    # Fila de total al final
    fila_total = len(ventas) + 2
    ws.cell(row=fila_total, column=6, value='TOTAL').font = Font(bold=True)
    celda_total = ws.cell(row=fila_total, column=7,
                        value=sum(float(v.total) for v in ventas))
    celda_total.font          = Font(bold=True, color=COLOR_HEADER)
    celda_total.number_format = '"$"#,##0.00'

    # Ancho de columnas
    anchos = [8, 18, 25, 18, 10, 18, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    # Guardar en buffer y devolver
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"ventas_{desde.isoformat()}_{hasta.isoformat()}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre_archivo,
    )


@admin_bp.route('/reportes/exportar/pdf')
@login_required
@admin_required
def exportar_pdf():
    from flask import send_file
    from fpdf import FPDF
    import io

    def _safe(s, maxlen=None):
        """Convierte a string Latin-1 seguro para fuentes core de fpdf2."""
        if not s:
            return '-'
        s = str(s)
        s = (s.replace('—', '-').replace('–', '-')
              .replace('“', '"').replace('”', '"')
              .replace('‘', "'").replace('’', "'")
              .replace('…', '...').replace('•', '-'))
        try:
            s.encode('latin-1')
        except UnicodeEncodeError:
            s = s.encode('latin-1', 'replace').decode('latin-1')
        return s[:maxlen] if maxlen else s

    desde_str   = request.args.get('desde', '')
    hasta_str   = request.args.get('hasta', '')
    metodo_pago = request.args.get('metodo_pago', '')

    hoy = date.today()
    try:
        desde = datetime.strptime(desde_str, '%Y-%m-%d').date() if desde_str else hoy.replace(day=1)
        hasta = datetime.strptime(hasta_str, '%Y-%m-%d').date() if hasta_str else hoy
    except ValueError:
        desde = hoy.replace(day=1)
        hasta = hoy

    query = Sale.query.filter(
        Sale.creado_en >= datetime.combine(desde, datetime.min.time()),
        Sale.creado_en <= datetime.combine(hasta, datetime.max.time()),
    )
    if metodo_pago:
        query = query.filter(Sale.metodo_pago == metodo_pago)
    ventas = query.order_by(Sale.creado_en.asc()).all()
    monto_total = sum(float(v.total or 0) for v in ventas)
    ticket_promedio = monto_total / len(ventas) if ventas else 0
    generado_en = datetime.now().strftime('%d/%m/%Y %H:%M')

    try:
        pdf = FPDF()
        pdf.set_margins(15, 15, 15)
        pdf.add_page()

        # ── Encabezado ────────────────────────────────────────────────
        pdf.set_fill_color(12, 12, 98)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_x(pdf.l_margin)
        pdf.cell(0, 14, 'DG Limpieza - Reporte de ventas', new_x='LMARGIN', new_y='NEXT', fill=True, align='L')

        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(80, 80, 80)
        periodo = f"Periodo: {desde.strftime('%d/%m/%Y')} - {hasta.strftime('%d/%m/%Y')}"
        if metodo_pago:
            periodo += f"  |  Metodo de pago: {_safe(metodo_pago)}"
        pdf.cell(0, 6, periodo, new_x='LMARGIN', new_y='NEXT')
        pdf.cell(0, 6, f'Generado: {generado_en}', new_x='LMARGIN', new_y='NEXT')
        pdf.ln(4)

        # ── Metricas ──────────────────────────────────────────────────
        pdf.set_fill_color(234, 246, 255)
        pdf.set_draw_color(57, 235, 225)
        pdf.set_line_width(0.8)
        col_w = (pdf.w - pdf.l_margin - pdf.r_margin) / 3

        metricas = [
            (str(len(ventas)), 'Ventas'),
            (f"${monto_total:,.2f}", 'Total facturado'),
            (f"${ticket_promedio:,.2f}", 'Ticket promedio'),
        ]
        x_start = pdf.l_margin
        y_start = pdf.get_y()
        for valor, etiqueta in metricas:
            pdf.set_xy(x_start, y_start)
            pdf.set_font('Helvetica', 'B', 14)
            pdf.set_text_color(12, 12, 98)
            pdf.cell(col_w - 2, 8, valor, border='LTR', fill=True, align='C', new_x='RIGHT', new_y='TOP')
            x_start += col_w

        x_start = pdf.l_margin
        for _, etiqueta in metricas:
            pdf.set_xy(x_start, y_start + 8)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(85, 85, 85)
            pdf.cell(col_w - 2, 6, etiqueta, border='LBR', fill=True, align='C', new_x='RIGHT', new_y='TOP')
            x_start += col_w

        pdf.ln(20)

        # ── Tabla ─────────────────────────────────────────────────────
        headers = ['#', 'Fecha', 'Cliente', 'Canal', 'Metodo pago', 'Total']
        widths  = [12, 36, 45, 22, 35, 30]

        pdf.set_fill_color(12, 12, 98)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_line_width(0.3)
        for h, w in zip(headers, widths):
            pdf.cell(w, 7, h, border=1, fill=True, align='C')
        pdf.ln()

        pdf.set_font('Helvetica', '', 9)
        for i, v in enumerate(ventas):
            fill = i % 2 == 0
            pdf.set_fill_color(234, 246, 255) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.set_text_color(26, 26, 46)
            row = [
                str(v.id),
                v.creado_en.strftime('%d/%m/%Y %H:%M') if v.creado_en else '-',
                _safe(v.nombre_cliente, 25),
                _safe(v.tipo or '').upper(),
                _safe(v.metodo_pago),
                f"${float(v.total or 0):,.2f}",
            ]
            for val, w in zip(row, widths):
                pdf.cell(w, 6, val, border=1, fill=fill)
            pdf.ln()

        # Fila totales
        pdf.set_fill_color(12, 12, 98)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 9)
        total_label_w = sum(widths[:-1])
        pdf.cell(total_label_w, 7, 'TOTAL', border=1, fill=True, align='R')
        pdf.cell(widths[-1], 7, f"${monto_total:,.2f}", border=1, fill=True, align='C')
        pdf.ln()

        # ── Pie ───────────────────────────────────────────────────────
        pdf.ln(6)
        pdf.set_font('Helvetica', '', 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 5, 'DG Limpieza - Reporte generado automaticamente', align='C')

        buffer = io.BytesIO(bytes(pdf.output()))
        nombre_archivo = f"reporte_{desde.isoformat()}_{hasta.isoformat()}.pdf"
        return send_file(buffer, mimetype='application/pdf',
                        as_attachment=True, download_name=nombre_archivo)

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'detalle': traceback.format_exc()}), 500

# ════════════════════════════════════════════════════════════════════════════
# COMBOS
# ════════════════════════════════════════════════════════════════════════════
from app.models import Combo, ComboItem

@admin_bp.route('/combos')
@login_required
@admin_required
def combos():
    todos = Combo.query.order_by(Combo.creado_en.desc()).all()
    return render_template('admin/combos.html', combos=todos)


@admin_bp.route('/combos/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def combo_nuevo():
    productos = Product.query.filter_by(activo=True).order_by(Product.nombre).all()

    if request.method == 'GET':
        return render_template('admin/combo_form.html',
            combo=None, productos=productos)

    nombre          = request.form.get('nombre', '').strip()
    descripcion     = request.form.get('descripcion', '').strip() or None
    imagen_url      = request.form.get('imagen_url', '').strip() or None
    precio_original = _precio_float(request.form.get('precio_original'))
    precio_combo    = _precio_float(request.form.get('precio_combo'))
    descuento_texto = request.form.get('descuento_texto', '').strip() or None

    if not nombre or not precio_combo:
        flash('El nombre y el precio del combo son obligatorios.', 'error')
        return render_template('admin/combo_form.html',
            combo=None, productos=productos)

    combo = Combo(
        nombre=nombre,
        descripcion=descripcion,
        imagen_url=imagen_url,
        precio_original=precio_original,
        precio_combo=precio_combo,
        descuento_texto=descuento_texto,
        activo=True
    )
    db.session.add(combo)
    db.session.flush()  # necesitamos combo.id antes de crear los items

    # Items del combo: listas paralelas producto_id[] y cantidad[]
    producto_ids = request.form.getlist('item_producto_id')
    cantidades   = request.form.getlist('item_cantidad')

    for pid, cant in zip(producto_ids, cantidades):
        if pid and cant:
            item = ComboItem(
                combo_id    = combo.id,
                producto_id = int(pid),
                cantidad    = int(cant) if cant else 1
            )
            db.session.add(item)

    db.session.commit()
    flash(f'Combo "{combo.nombre}" creado correctamente.', 'success')
    return redirect(url_for('admin.combos'))


@admin_bp.route('/combos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@admin_required
def combo_editar(id):
    combo     = Combo.query.get_or_404(id)
    productos = Product.query.filter_by(activo=True).order_by(Product.nombre).all()

    if request.method == 'GET':
        return render_template('admin/combo_form.html',
            combo=combo, productos=productos)

    combo.nombre          = request.form.get('nombre', '').strip()
    combo.descripcion     = request.form.get('descripcion', '').strip() or None
    combo.imagen_url      = request.form.get('imagen_url', '').strip() or None
    combo.precio_original = _precio_float(request.form.get('precio_original'))
    combo.precio_combo    = _precio_float(request.form.get('precio_combo'))
    combo.descuento_texto = request.form.get('descuento_texto', '').strip() or None

    if not combo.nombre or not combo.precio_combo:
        flash('El nombre y el precio del combo son obligatorios.', 'error')
        return render_template('admin/combo_form.html',
            combo=combo, productos=productos)

    # Reemplazar items: borrar los anteriores y recrear
    ComboItem.query.filter_by(combo_id=combo.id).delete()

    producto_ids = request.form.getlist('item_producto_id')
    cantidades   = request.form.getlist('item_cantidad')

    for pid, cant in zip(producto_ids, cantidades):
        if pid and cant:
            item = ComboItem(
                combo_id    = combo.id,
                producto_id = int(pid),
                cantidad    = int(cant) if cant else 1
            )
            db.session.add(item)

    db.session.commit()
    flash(f'Combo "{combo.nombre}" actualizado.', 'success')
    return redirect(url_for('admin.combos'))


@admin_bp.route('/combos/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def combo_toggle(id):
    combo = Combo.query.get_or_404(id)
    combo.activo = not combo.activo
    db.session.commit()
    estado = 'activado' if combo.activo else 'desactivado'
    flash(f'Combo "{combo.nombre}" {estado}.', 'success')
    return redirect(url_for('admin.combos'))

# ════════════════════════════════════════════════════════════════════════════
# PEDIDOS WHATSAPP PENDIENTES
# ════════════════════════════════════════════════════════════════════════════

@admin_bp.route('/pedidos')
@login_required
@admin_required
def pedidos():
    """
    Lista los pedidos WhatsApp con estado 'pendiente'.
    Son los pedidos que el cliente generó desde la tienda pero
    todavía no fueron confirmados ni cobrados por el vendedor.
    """
    pedidos_list = (
        Sale.query
        .filter_by(tipo='whatsapp', estado='pendiente')
        .order_by(Sale.creado_en.desc())
        .all()
    )
    return render_template('admin/pedidos.html', pedidos=pedidos_list)


@admin_bp.route('/pedidos/<int:id>/confirmar', methods=['GET', 'POST'])
@login_required
@admin_required
def pedido_confirmar(id):
    """
    GET  → muestra el detalle del pedido con formulario para elegir método de pago.
    POST → confirma el pedido: descuenta stock, cambia estado a 'completada',
        asigna el vendedor que lo procesó.

    ¿Por qué descontamos stock recién acá y no cuando el cliente confirma?
    Porque hasta este momento el pedido puede cancelarse. El stock se reserva
    en el mundo real cuando el vendedor lo confirma y prepara el paquete.
    """
    pedido = Sale.query.get_or_404(id)

    if pedido.estado != 'pendiente':
        flash('Este pedido ya fue procesado.', 'info')
        return redirect(url_for('admin.pedidos'))

    if request.method == 'GET':
        return render_template('admin/pedido_confirmar.html', pedido=pedido)

    # ── POST: confirmar el pedido ─────────────────────────────────────────
    metodo_pago = request.form.get('metodo_pago', '').strip()
    if not metodo_pago:
        flash('Seleccioná un método de pago.', 'error')
        return render_template('admin/pedido_confirmar.html', pedido=pedido)

    # Descontar stock por cada item del pedido
    for item in pedido.items:
        if item.combo_id:
            # Para combos descontamos cada producto componente
            combo_items = ComboItem.query.filter_by(combo_id=item.combo_id).all()
            for ci in combo_items:
                producto = Product.query.get(ci.producto_id)
                if producto and producto.stock is not None:
                    descuento = ci.cantidad * item.cantidad
                    producto.stock = max(0, producto.stock - descuento)
                    db.session.add(StockMovement(
                        producto_id = producto.id,
                        usuario_id  = current_user.id,
                        tipo        = 'salida',
                        cantidad    = descuento,
                        motivo      = f'Pedido WhatsApp #{pedido.id} — combo "{item.nombre_producto}"'
                    ))
        else:
            producto = Product.query.get(item.producto_id)
            if producto and producto.stock is not None:
                producto.stock = max(0, producto.stock - item.cantidad)
                db.session.add(StockMovement(
                    producto_id = producto.id,
                    usuario_id  = current_user.id,
                    tipo        = 'salida',
                    cantidad    = item.cantidad,
                    motivo      = f'Pedido WhatsApp #{pedido.id} confirmado'
                ))

    # Actualizar el pedido
    pedido.estado      = 'completada'
    pedido.metodo_pago = metodo_pago
    pedido.usuario_id  = current_user.id  # asignamos al vendedor que lo procesó

    db.session.commit()
    flash(f'Pedido #{pedido.id} confirmado correctamente.', 'success')
    return redirect(url_for('admin.pedidos'))


@admin_bp.route('/pedidos/<int:id>/cancelar', methods=['POST'])
@login_required
@admin_required
def pedido_cancelar(id):
    """
    Cancela un pedido pendiente sin descontar stock.
    Útil cuando el cliente no vino a retirar o canceló por WhatsApp.
    """
    pedido = Sale.query.get_or_404(id)

    if pedido.estado != 'pendiente':
        flash('Este pedido ya fue procesado.', 'info')
        return redirect(url_for('admin.pedidos'))

    pedido.estado = 'cancelada'
    db.session.commit()
    flash(f'Pedido #{pedido.id} cancelado.', 'info')
    return redirect(url_for('admin.pedidos'))

# ════════════════════════════════════════════════════════
# UPLOAD DE IMAGEN A CLOUDINARY
# ════════════════════════════════════════════════════════
@admin_bp.route('/upload-imagen', methods=['POST'])
@login_required
@admin_required
def upload_imagen():
    """
    Recibe una imagen desde el formulario de producto,
    la sube a Cloudinary y devuelve la URL segura.
    
    ¿Por qué una ruta separada y no procesar en producto_nuevo?
    Porque así el upload es independiente del guardado del producto.
    Si el usuario cambia la imagen varias veces antes de guardar,
    solo hacemos un upload, no mezclamos lógicas.
    """
    import cloudinary.uploader

    if 'imagen' not in request.files:
        return jsonify({'error': 'No se recibió ninguna imagen'}), 400

    archivo = request.files['imagen']

    if archivo.filename == '':
        return jsonify({'error': 'Archivo vacío'}), 400

    # Validar que sea una imagen
    extensiones_permitidas = {'jpg', 'jpeg', 'png', 'webp', 'gif'}
    extension = archivo.filename.rsplit('.', 1)[-1].lower()
    if extension not in extensiones_permitidas:
        return jsonify({'error': 'Formato no permitido. Usá JPG, PNG o WEBP'}), 400

    try:
        resultado = cloudinary.uploader.upload(
            archivo,
            folder='dg-limpieza/productos',   # carpeta dentro de tu cuenta
            transformation=[
                {'width': 800, 'height': 800, 'crop': 'limit'},  # máximo 800x800
                {'quality': 'auto'},                               # optimiza el peso
                {'fetch_format': 'auto'}                           # convierte a WebP si el browser lo soporta
            ]
        )
        return jsonify({'url': resultado['secure_url']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500